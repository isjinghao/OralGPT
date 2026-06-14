#!/usr/bin/env python3
"""
Build one integrated diagnosis QA JSON per patient.

Inputs:
  1) textrecord_*.xlsx in each patient folder, especially the Diagnosis sheet.
  2) Per-patient JSON records from examination, XLa/CT, and panoramic report pipelines.

Output:
  One JSON per patient, compatible with the existing per-patient JSON shape.
"""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import os
import re
import unicodedata
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openai import OpenAI
from openpyxl import load_workbook


QUESTION_DIAGNOSIS = (
    "What is the diagnosis based on the clinical, radiological, and functional findings?"
)

DEFAULT_PROMPT_TEMPLATE = """
You are an expert in orthodontics, orthognathic surgery, and oral and maxillofacial surgery.

Tasks:
1. Integrate all available imaging information into coherent diagnostic conclusions, rather than listing isolated observations.
2. Use standardized medical terminology consistent with orthodontics and orthognathic surgery.
3. Clearly distinguish between:
    * Dentofacial deformity-related diagnoses (for surgical/orthodontic planning)
    * General dental and systemic disease diagnoses (for comprehensive treatment planning)
4. Avoid vague language; provide definitive diagnosis whenever supported by imaging.
5. If a category cannot be assessed, explicitly state: “Not assessable based on current data.”

Output Format:

Part I: Dentofacial Deformity-Related Diagnosis

(Focus on deformity classification and surgical relevance)

1. Skeletal Malocclusion (Orthognathic Perspective)

* Sagittal skeletal pattern (Class I / II / III; specify dominant component, e.g., mandibular prognathism or retrognathia, maxillary deficiency/excess)
* Vertical pattern (hyperdivergent / hypodivergent / normodivergent)
* Transverse relationship (facial asymmetry, occlusal cant, yaw deformity if present)

2. Dental Malocclusion (Orthodontic Perspective)

* Angle classification (Class I / II / III, if inferable)
* Overjet / overbite condition (e.g., crossbite, open bite, deep bite)
* Dental compensation (incisor proclination/retroclination)
* Alignment (crowding/spacing, if inferable)

3. Craniofacial and Structural Conditions

* Presence of congenital syndromes (if suggested)
* Condylar pathology (e.g., condylar osteochondroma, hyperplasia, hypoplasia)
* Jaw defects (e.g., maxillary/mandibular deficiency, post-surgical or pathological defects)
* Airway-related conditions (e.g., obstructive sleep apnea, if inferable)
* Temporomandibular joint disorders (if signs present)

4. Integrated Dentofacial Diagnosis

* Provide a concise summary sentence integrating skeletal, dental, and facial features
    (e.g., “Skeletal Class III dentofacial deformity characterized by mandibular prognathism with maxillary deficiency and hypodivergent pattern, without significant asymmetry.”)

Part II: Dental and Systemic Disease Diagnosis

(Focus on general oral and systemic health relevant to treatment)

1. Dental Hard Tissue Diseases

* Caries, tooth defects, non-carious lesions (if assessable)

2. Periodontal Conditions

* Periodontal status (e.g., gingivitis, periodontitis, bone loss if visible)

3. Dentition and Prosthodontic Status

* Missing teeth, impacted teeth, restorations, prostheses (if assessable)

4. Other Oral Pathologies

* Cysts, tumors, or other lesions (if present)

5. Systemic Conditions

* Any systemic diseases relevant to treatment (if provided or inferable)

Style Requirements:

* Use clear, structured, and professional clinical language
* Ensure output resembles a formal diagnosis sheet used in orthognathic treatment planning or clinical documentation
* Keep statements precise, non-redundant, and decision-oriented

Question: {question}

Diagnosis sheet content:
{diagnosis_text}

Clinical, radiological, and functional QA pairs extracted from JSON Modalities:
{findings_qa_text}
""".strip()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Use an OpenAI-compatible LLM to integrate diagnosis and multimodal findings into one QA pair."
    )
    parser.add_argument(
        "--dataset-root",
        default="/data/OralGPT/OralGPT-CMF/dataset/SH9HCMFdata",
        help="Path to SH9HCMFdata root.",
    )
    parser.add_argument(
        "--output-dir",
        default="./outputs/stage3_patient_json_diagnosis_qa",
        help="Directory to save per-patient diagnosis QA JSON files.",
    )
    parser.add_argument(
        "--examination-json-dir",
        default="./outputs/stage1_patient_json_examination",
        help="Directory containing stage1 examination JSON files.",
    )
    parser.add_argument(
        "--xla-ct-json-dir",
        default="./outputs/stage2_patient_json_XLa_CT",
        help="Directory containing stage2 XLa/CT JSON files.",
    )
    parser.add_argument(
        "--xray-report-json-dir",
        default="./outputs/stage2_patient_json_xray_oralagent_report",
        help="Directory containing stage2 xray/oralagent report JSON files.",
    )
    parser.add_argument(
        "--model",
        default="gpt-4o",
        help="OpenAI model name for diagnosis synthesis.",
    )
    parser.add_argument(
        "--api-base",
        default="http://35.164.11.19:3887/v1",
        help="OpenAI compatible API base URL.",
    )
    parser.add_argument(
        "--api-key-env",
        default="OPENAI_API_KEY",
        help="Environment variable name that stores API key.",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Overwrite existing patient JSON files.",
    )
    parser.add_argument(
        "--max-patients",
        type=int,
        default=None,
        help="Only process first N patients for debugging.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Do not call LLM; write deterministic fallback answers.",
    )
    parser.add_argument(
        "--workers",
        type=int,
        default=1,
        help="Number of worker threads for patient-level parallel processing.",
    )
    return parser.parse_args()


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value)).strip()
    if not text:
        return ""
    return re.sub(r"\s+", " ", text)


def discover_patients(dataset_root: Path) -> List[Dict[str, Any]]:
    if not dataset_root.exists():
        raise FileNotFoundError(f"Dataset root does not exist: {dataset_root}")

    patients: List[Dict[str, Any]] = []
    for group_dir in sorted(p for p in dataset_root.iterdir() if p.is_dir() and p.name.startswith("group")):
        for patient_dir in sorted(p for p in group_dir.iterdir() if p.is_dir()):
            textrecords = sorted(patient_dir.glob("textrecord_*.xlsx"))
            patients.append(
                {
                    "group": group_dir.name,
                    "patient_name": patient_dir.name,
                    "patient_dir": patient_dir,
                    "textrecord": textrecords[0] if textrecords else None,
                }
            )
    return patients


def build_unique_patient_uid(
    group: str, patient_name: str, patient_dir: Path, used: Dict[str, int]
) -> str:
    base = f"{group}__{patient_name}"
    if base not in used:
        used[base] = 1
        return base
    used[base] += 1
    short_hash = hashlib.sha1(str(patient_dir).encode("utf-8")).hexdigest()[:8]
    return f"{base}__{short_hash}"


def json_record_template(
    uid: str,
    group: str,
    patient_name: str,
    patient_dir: Path,
    model: str,
) -> Dict[str, Any]:
    now = dt.datetime.now(dt.timezone.utc).isoformat()
    return {
        "schema_version": "1.0",
        "patient_uid": uid,
        "patient_name": patient_name,
        "group": group,
        "source_path": str(patient_dir),
        "created_at_utc": now,
        "updated_at_utc": now,
        "meta": {
            "model": model,
        },
        "Modalities": {},
    }


def get_sheet(workbook: Any, target_name: str) -> Optional[Any]:
    target = normalize_text(target_name).lower()
    for sheet_name in workbook.sheetnames:
        if normalize_text(sheet_name).lower() == target:
            return workbook[sheet_name]
    for sheet_name in workbook.sheetnames:
        if target in normalize_text(sheet_name).lower():
            return workbook[sheet_name]
    return None


def iter_non_empty_rows(sheet: Any) -> Iterable[List[str]]:
    for row in sheet.iter_rows(values_only=True):
        values = [normalize_text(cell) for cell in row]
        if any(values):
            yield [v for v in values if v]


def read_diagnosis_sheet(textrecord_path: Optional[Path]) -> Tuple[List[str], str]:
    if textrecord_path is None or not textrecord_path.exists():
        return [], "Missing textrecord_*.xlsx."

    workbook = load_workbook(filename=str(textrecord_path), data_only=True, read_only=True)
    sheet = get_sheet(workbook, "Diagnosis")
    if sheet is None:
        return [], "Missing Diagnosis sheet."

    entries: List[str] = []
    for row in iter_non_empty_rows(sheet):
        row_text = "; ".join(row)
        if row_text:
            entries.append(row_text)
    return entries, ""


def read_json(path: Path) -> Optional[Dict[str, Any]]:
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None


def collect_modality_findings(uid: str, json_dirs: Dict[str, Path]) -> Tuple[List[Dict[str, str]], List[str]]:
    findings: List[Dict[str, str]] = []
    missing_or_failed: List[str] = []

    for source_name, json_dir in json_dirs.items():
        json_path = json_dir / f"{uid}.json"
        data = read_json(json_path)
        if data is None:
            missing_or_failed.append(f"{source_name}: missing or unreadable JSON")
            continue

        modalities = data.get("Modalities", {})
        if not isinstance(modalities, dict):
            missing_or_failed.append(f"{source_name}: missing Modalities object")
            continue

        for modality_name, modality in sorted(modalities.items()):
            if not isinstance(modality, dict):
                continue
            answer = normalize_text(modality.get("answer", ""))
            if not answer:
                continue
            findings.append(
                {
                    "source": source_name,
                    "modality": str(modality_name),
                    "question": normalize_text(modality.get("question", "")),
                    "answer": answer,
                }
            )

    return findings, missing_or_failed


def format_findings_qa_text(findings: List[Dict[str, str]]) -> str:
    if not findings:
        return "Not available."

    blocks: List[str] = []
    for item in findings:
        source = item.get("source", "unknown_source")
        modality = item.get("modality", "unknown_modality")
        question = item.get("question", "")
        answer = item.get("answer", "")
        if not answer:
            continue
        blocks.append(
            "\n".join(
                [
                    f"[{source} / {modality}]",
                    f"Question: {question if question else 'Not available.'}",
                    f"Answer: {answer}",
                ]
            )
        )
    return "\n\n".join(blocks) if blocks else "Not available."


def fallback_answer(diagnosis_entries: List[str], findings: List[Dict[str, str]]) -> str:
    diagnosis = "; ".join(diagnosis_entries).strip()
    support_parts = [
        f"{item['modality']}: {item['answer']}"
        for item in findings
        if item.get("answer")
    ]
    support = " ".join(support_parts)
    if diagnosis and support:
        return f"The diagnosis is {diagnosis}, supported by the following clinical, radiological, and functional findings: {support}"
    if diagnosis:
        return f"The diagnosis is {diagnosis}."
    if support:
        return f"The diagnosis should be inferred from the available findings: {support}"
    return ""


def call_llm_diagnosis(
    client: OpenAI,
    model: str,
    diagnosis_entries: List[str],
    findings: List[Dict[str, str]],
) -> str:
    diagnosis_text = "; ".join(diagnosis_entries) if diagnosis_entries else "Not available."
    prompt = DEFAULT_PROMPT_TEMPLATE.format(
        question=QUESTION_DIAGNOSIS,
        diagnosis_text=diagnosis_text,
        findings_qa_text=format_findings_qa_text(findings),
    )
    completion = client.chat.completions.create(
        model=model,
        messages=[
            {
                "role": "system",
                "content": "You produce concise, faithful, clinically precise English diagnosis summaries.",
            },
            {"role": "user", "content": prompt},
        ],
    )
    return normalize_text(completion.choices[0].message.content or "")


def add_diagnosis_task(
    record: Dict[str, Any],
    textrecord_path: Optional[Path],
    json_paths: Dict[str, Path],
    diagnosis_entries: List[str],
    findings: List[Dict[str, str]],
    source_warnings: List[str],
    status: str,
    answer: str = "",
    error: str = "",
) -> None:
    record["Modalities"]["integrated_diagnosis"] = {
        "question": QUESTION_DIAGNOSIS,
        "status": status,
        "input_images": [],
        "input_textrecord": str(textrecord_path) if textrecord_path else "",
        "input_jsons": {name: str(path) for name, path in json_paths.items()},
        "diagnosis_sheet_entries": diagnosis_entries,
        "source_qa_pairs": findings,
        "source_warnings": source_warnings,
        "num_integrated_findings": len(findings),
        "answer": answer,
        "error": error,
        "timestamp_utc": dt.datetime.now(dt.timezone.utc).isoformat(),
    }
    record["updated_at_utc"] = dt.datetime.now(dt.timezone.utc).isoformat()


def process_one_patient(
    item: Dict[str, Any],
    uid: str,
    output_path: Path,
    args: argparse.Namespace,
    client: OpenAI | None,
    json_dirs: Dict[str, Path],
) -> Dict[str, Any]:
    if output_path.exists() and not args.overwrite:
        return {
            "uid": uid,
            "output_path": output_path,
            "skipped": True,
            "status": "exists",
        }

    record = json_record_template(
        uid=uid,
        group=item["group"],
        patient_name=item["patient_name"],
        patient_dir=item["patient_dir"],
        model=args.model if client else "fallback_rule_based",
    )

    diagnosis_entries, diagnosis_warning = read_diagnosis_sheet(item["textrecord"])
    findings, source_warnings = collect_modality_findings(uid, json_dirs)
    if diagnosis_warning:
        source_warnings.append(f"Diagnosis sheet: {diagnosis_warning}")

    json_paths = {name: path / f"{uid}.json" for name, path in json_dirs.items()}
    fallback = fallback_answer(diagnosis_entries, findings)

    if args.dry_run:
        add_diagnosis_task(
            record=record,
            textrecord_path=item["textrecord"],
            json_paths=json_paths,
            diagnosis_entries=diagnosis_entries,
            findings=findings,
            source_warnings=source_warnings,
            status="dry_run",
            answer=fallback,
        )
    elif not diagnosis_entries and not findings:
        add_diagnosis_task(
            record=record,
            textrecord_path=item["textrecord"],
            json_paths=json_paths,
            diagnosis_entries=diagnosis_entries,
            findings=findings,
            source_warnings=source_warnings,
            status="missing_input",
            error="No diagnosis sheet entries or modality findings were available.",
        )
    else:
        try:
            assert client is not None
            answer = call_llm_diagnosis(
                client=client,
                model=args.model,
                diagnosis_entries=diagnosis_entries,
                findings=findings,
            )
            add_diagnosis_task(
                record=record,
                textrecord_path=item["textrecord"],
                json_paths=json_paths,
                diagnosis_entries=diagnosis_entries,
                findings=findings,
                source_warnings=source_warnings,
                status="success",
                answer=answer if answer else fallback,
            )
        except Exception as exc:
            add_diagnosis_task(
                record=record,
                textrecord_path=item["textrecord"],
                json_paths=json_paths,
                diagnosis_entries=diagnosis_entries,
                findings=findings,
                source_warnings=source_warnings,
                status="failed",
                answer=fallback,
                error=str(exc),
            )

    output_path.write_text(json.dumps(record, ensure_ascii=False, indent=2), encoding="utf-8")
    status = record["Modalities"]["integrated_diagnosis"]["status"]
    return {
        "uid": uid,
        "output_path": output_path,
        "skipped": False,
        "status": status,
    }


def run_pipeline(args: argparse.Namespace) -> None:
    dataset_root = Path(args.dataset_root).resolve()
    output_dir = Path(args.output_dir).resolve()
    ensure_dir(output_dir)

    base_dir = Path(__file__).resolve().parent
    json_dirs = {
        "examination": (base_dir / args.examination_json_dir).resolve()
        if not Path(args.examination_json_dir).is_absolute()
        else Path(args.examination_json_dir).resolve(),
        "xla_ct": (base_dir / args.xla_ct_json_dir).resolve()
        if not Path(args.xla_ct_json_dir).is_absolute()
        else Path(args.xla_ct_json_dir).resolve(),
        "xray_oralagent_report": (base_dir / args.xray_report_json_dir).resolve()
        if not Path(args.xray_report_json_dir).is_absolute()
        else Path(args.xray_report_json_dir).resolve(),
    }

    api_key = os.getenv(args.api_key_env, "")
    if (not args.dry_run) and (not api_key):
        raise RuntimeError(f"Missing API key. Set environment variable: {args.api_key_env}")
    client = None if args.dry_run else OpenAI(api_key=api_key, base_url=args.api_base)

    patients = discover_patients(dataset_root)
    if args.max_patients is not None:
        patients = patients[: args.max_patients]

    used_uid: Dict[str, int] = {}
    jobs: List[Tuple[Dict[str, Any], str, Path]] = []
    for item in patients:
        uid = build_unique_patient_uid(item["group"], item["patient_name"], item["patient_dir"], used_uid)
        jobs.append((item, uid, output_dir / f"{uid}.json"))

    total = len(jobs)
    workers = max(1, args.workers)
    print(
        f"[INFO] Found {total} patients under {dataset_root}; "
        f"workers={workers}, dry_run={args.dry_run}"
    )

    completed = 0
    with ThreadPoolExecutor(max_workers=workers) as executor:
        future_map = {
            executor.submit(process_one_patient, item, uid, out, args, client, json_dirs): (uid, out)
            for item, uid, out in jobs
        }
        for future in as_completed(future_map):
            uid, out = future_map[future]
            completed += 1
            progress = (completed / total * 100.0) if total else 100.0
            try:
                result = future.result()
                if result["skipped"]:
                    print(
                        f"[{completed}/{total} {progress:.1f}%] "
                        f"[SKIP] {uid}: output exists ({out})"
                    )
                else:
                    print(
                        f"[{completed}/{total} {progress:.1f}%] "
                        f"[DONE] {uid} -> {out} | integrated_diagnosis:{result['status']}"
                    )
            except Exception as exc:
                print(f"[{completed}/{total} {progress:.1f}%] [ERROR] {uid}: {exc}")

    print("[INFO] Pipeline completed.")


if __name__ == "__main__":
    run_pipeline(parse_args())
