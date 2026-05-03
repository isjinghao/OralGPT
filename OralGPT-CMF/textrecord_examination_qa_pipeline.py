#!/usr/bin/env python3
"""
Build per-patient QA JSON files from textrecord_*.xlsx.

Expected dataset layout (same style as existing CMF pipeline):
  dataset_root/
    group1..groupN/
      PATIENT_NAME/
        textrecord_*.xlsx

Output:
  One JSON per patient, compatible with existing per-patient JSON shape.
"""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import os
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openai import OpenAI
from openpyxl import load_workbook


QUESTION_BASIC_INFO = "Please provide your basic information, including age, gender, height, etc."
QUESTION_PRIMARY_CONCERN = "What is the primary concern of the patient?"
QUESTION_PAST_HISTORY = "What is the patient’s past medical or surgical history?"
QUESTION_FACIAL = "What objective findings can be observed from the facial photographs?"
QUESTION_DENTAL = "What are the occlusal and dental findings based on the intraoral examination?"
QUESTION_TMJ = "What are the temporomandibular joint (TMJ) findings on clinical examination?"
QUESTION_ECT = (
    "Was single-photon emission computed tomography (ECT) performed? "
    "If so, is the condyle stable?"
)
DOB_KEYWORDS = ["dob", "date of birth", "birth date", "birthday"]
AGE_REFERENCE_YEAR = 2026


def parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(description="Build QA JSON from textrecord_*.xlsx.")
    ap.add_argument(
        "--dataset-root",
        default="/data/OralGPT/OralGPT-CMF/dataset/SH9HCMFdata",
        help="Dataset root containing group*/patient directories.",
    )
    ap.add_argument(
        "--output-dir",
        default="./outputs/patient_json_textrecord",
        help="Directory to save generated per-patient JSON files.",
    )
    ap.add_argument(
        "--overwrite",
        action="store_true",
        help="Overwrite existing patient JSON files.",
    )
    ap.add_argument(
        "--max-patients",
        type=int,
        default=None,
        help="Only process first N patients for debugging.",
    )
    ap.add_argument(
        "--workers",
        type=int,
        default=1,
        help="Number of worker threads for patient-level parallel processing.",
    )
    ap.add_argument(
        "--model",
        default="gpt-4o",
        help="OpenAI model used to synthesize polished English answers.",
    )
    ap.add_argument(
        "--api-base",
        default="http://35.164.11.19:3887/v1",
        help="OpenAI compatible API base URL.",
    )
    ap.add_argument(
        "--api-key-env",
        default="OPENAI_API_KEY",
        help="Environment variable name that stores API key.",
    )
    ap.add_argument(
        "--dry-run",
        action="store_true",
        help="Do not call LLM; use deterministic fallback sentence assembly.",
    )
    return ap.parse_args()


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def normalize_text(s: Any) -> str:
    if s is None:
        return ""
    text = str(s).strip()
    if not text:
        return ""
    text = re.sub(r"\s+", " ", text)
    return text


def normalize_key(s: str) -> str:
    s = normalize_text(s).lower()
    s = s.replace("_", " ")
    s = re.sub(r"[^a-z0-9]+", "", s)
    return s


def discover_patients(dataset_root: Path) -> List[Dict[str, Any]]:
    if not dataset_root.exists():
        raise FileNotFoundError(f"Dataset root does not exist: {dataset_root}")

    items: List[Dict[str, Any]] = []
    for group_dir in sorted(p for p in dataset_root.iterdir() if p.is_dir() and p.name.startswith("group")):
        for patient_dir in sorted(p for p in group_dir.iterdir() if p.is_dir()):
            textrecords = sorted(patient_dir.glob("textrecord_*.xlsx"))
            items.append(
                {
                    "group": group_dir.name,
                    "patient_name": patient_dir.name,
                    "patient_dir": patient_dir,
                    "textrecord": textrecords[0] if textrecords else None,
                }
            )
    return items


def build_unique_patient_uid(group: str, patient_name: str, patient_dir: Path, used: Dict[str, int]) -> str:
    base = f"{group}__{patient_name}"
    if base not in used:
        used[base] = 1
        return base
    used[base] += 1
    short_hash = hashlib.sha1(str(patient_dir).encode("utf-8")).hexdigest()[:8]
    return f"{base}__{short_hash}"


def json_record_template(uid: str, group: str, patient_name: str, patient_dir: Path) -> Dict[str, Any]:
    now = dt.datetime.now(dt.timezone.utc).isoformat()
    return {
        "schema_version": "1.0",
        "patient_uid": uid,
        "patient_name": patient_name,
        "group": group,
        "source_path": str(patient_dir),
        "created_at_utc": now,
        "updated_at_utc": now,
        "meta": {
            "model": "rule_based_textrecord_pipeline",
        },
        "Modalities": {},
    }


def add_task_result(
    record: Dict[str, Any],
    task_type: str,
    question: str,
    textrecord_path: Path,
    answer: str,
) -> None:
    record["Modalities"][task_type] = {
        "question": question,
        "status": "success",
        "input_images": [],
        "input_textrecord": str(textrecord_path),
        "answer": answer,
        "error": "",
        "timestamp_utc": dt.datetime.now(dt.timezone.utc).isoformat(),
    }
    record["updated_at_utc"] = dt.datetime.now(dt.timezone.utc).isoformat()


def get_sheet(workbook: Any, target_name: str) -> Optional[Any]:
    by_norm = {normalize_text(name).lower(): name for name in workbook.sheetnames}
    key = normalize_text(target_name).lower()
    if key in by_norm:
        return workbook[by_norm[key]]
    # Fuzzy fallback for small naming variations.
    for n in workbook.sheetnames:
        if key in normalize_text(n).lower():
            return workbook[n]
    return None


def sheet_rows(sheet: Any) -> Iterable[List[str]]:
    for row in sheet.iter_rows(values_only=True):
        values = [normalize_text(c) for c in row]
        if any(values):
            yield values


def row_to_pair(values: List[str]) -> Optional[Tuple[str, str]]:
    non_empty = [v for v in values if v]
    if len(non_empty) < 2:
        return None
    key = non_empty[0]
    value = " ".join(non_empty[1:]).strip(" :;,.")
    if key and value:
        return key, value
    return None


def sheet_to_pairs(sheet: Any) -> List[Tuple[str, str]]:
    """
    Convert a sheet into (field, value) pairs with robust heuristics.
    """
    pairs: List[Tuple[str, str]] = []
    for row in sheet_rows(sheet):
        pair = row_to_pair(row)
        if pair is not None:
            pairs.append(pair)
    return pairs


def is_mri_row(values: List[str]) -> bool:
    return any("mri" in normalize_text(v).lower() for v in values)


def is_ect_key(value: str) -> bool:
    return re.search(r"(?<![a-z0-9])ect(?![a-z0-9])", normalize_text(value).lower()) is not None


def extract_numbers(text: str) -> List[float]:
    numbers: List[float] = []
    for match in re.findall(r"[-+]?(?:\d+(?:\.\d*)?|\.\d+)", text):
        try:
            numbers.append(float(match))
        except ValueError:
            continue
    return numbers


def tmj_sheet_to_pairs_without_mri(sheet: Any) -> List[Tuple[str, str]]:
    """
    Convert TMJ sheet rows to pairs while excluding MRI rows and their continuation rows.

    Some TMJ sheets store MRI details as a labeled MRI row followed by rows whose first
    column is empty. Those continuation rows still belong to the MRI section.
    """
    pairs: List[Tuple[str, str]] = []
    in_mri_block = False

    for raw_row in sheet.iter_rows(values_only=True):
        values = [normalize_text(c) for c in raw_row]
        if not any(values):
            in_mri_block = False
            continue

        first_cell = values[0] if values else ""
        row_is_mri = is_mri_row(values)
        if row_is_mri or (in_mri_block and not first_cell):
            in_mri_block = True
            continue

        in_mri_block = False
        pair = row_to_pair(values)
        if pair is not None:
            pairs.append(pair)
    return pairs


def build_ect_answer(sheet: Any) -> str:
    for raw_row in sheet.iter_rows(values_only=True):
        values = [normalize_text(c) for c in raw_row]
        if not values or not is_ect_key(values[0]):
            continue

        result_text = " ".join(v for v in values[1:] if v)
        numbers = extract_numbers(result_text)
        if not numbers:
            return (
                "Yes, single-photon emission computed tomography (ECT) was performed, "
                f"but condylar stability could not be determined from the recorded value: {result_text}."
            )

        is_unstable = any(abs(n) > 0.1 for n in numbers)
        stability = "unstable" if is_unstable else "stable"
        return (
            "Yes, single-photon emission computed tomography (ECT) was performed. "
            f"The condyle is {stability} based on the recorded ECT value of {result_text}."
        )
    return ""


def pairs_to_compact_json(pairs: List[Tuple[str, str]]) -> str:
    payload = [{"field": k, "value": v} for k, v in pairs if normalize_text(k) and normalize_text(v)]
    return json.dumps(payload, ensure_ascii=False)


def extract_field(pairs: List[Tuple[str, str]], keywords: List[str]) -> str:
    norm_keywords = [normalize_key(k) for k in keywords]
    for k, v in pairs:
        nk = normalize_key(k)
        if any(kw in nk for kw in norm_keywords):
            return normalize_text(v)
    # Fallback: check if keyword appears inside raw key.
    for k, v in pairs:
        lk = normalize_text(k).lower()
        if any(kw.lower() in lk for kw in keywords):
            return normalize_text(v)
    return ""


def build_basic_info_answers(sheet: Any) -> Dict[str, str]:
    pairs = sheet_to_pairs(sheet)
    age = extract_field(pairs, DOB_KEYWORDS)
    gender = extract_field(pairs, ["gender", "sex"])
    height = extract_field(pairs, ["height"])
    weight = extract_field(pairs, ["weight"])
    cc = extract_field(pairs, ["cc", "chief complaint", "chiefconcern", "primary concern"])
    rmh = extract_field(
        pairs,
        ["rmh", "medical history", "surgical history", "past medical", "past surgical"],
    )

    answers: Dict[str, str] = {}

    basic_parts: List[str] = []
    if age:
        basic_parts.append(f"The recorded DOB or age value is {age}.")
    if gender:
        basic_parts.append(f"The patient is {gender}.")
    if height:
        basic_parts.append(f"The height is {height}.")
    if weight:
        basic_parts.append(f"The weight is {weight}.")
    if basic_parts:
        answers["basic_info"] = " ".join(basic_parts)

    if cc:
        answers["primary_concern"] = f"The patient's primary concern is {cc}."
    if rmh:
        answers["past_history"] = f"The patient's past medical or surgical history includes {rmh}."
    return answers


def pairs_to_paragraph(pairs: List[Tuple[str, str]]) -> str:
    clauses: List[str] = []
    for k, v in pairs:
        key = normalize_text(k).strip(" :")
        val = normalize_text(v)
        if not key or not val:
            continue
        clauses.append(f"{key} is {val}")
    if not clauses:
        return ""
    return ". ".join(clauses) + "."


def build_facial_answer(sheet: Any) -> str:
    pairs = sheet_to_pairs(sheet)
    return pairs_to_paragraph(pairs)


def build_dental_answer(sheet: Any) -> str:
    pairs = sheet_to_pairs(sheet)
    return pairs_to_paragraph(pairs)


def build_tmj_answer(sheet: Any) -> str:
    filtered_pairs = tmj_sheet_to_pairs_without_mri(sheet)
    return pairs_to_paragraph(filtered_pairs)


def call_llm_answer(
    client: OpenAI,
    model: str,
    question: str,
    sheet_name: str,
    pairs: List[Tuple[str, str]],
) -> str:
    prompt = (
        "You are a medical writing assistant for structured clinical spreadsheet data.\n"
        "Task:\n"
        "- Read the structured key-value data from one spreadsheet sheet.\n"
        "- Answer the question in natural, fluent English.\n"
        "- Keep all clinically meaningful details.\n"
        "- Do NOT invent values.\n"
        "- If some details are missing, ignore them silently.\n"
        f"- For basic information, if DOB is a numeric age, state it as the age; "
        f"if DOB is a date of birth, calculate the age using {AGE_REFERENCE_YEAR} as the reference year.\n"
        "- Output exactly one paragraph and nothing else.\n\n"
        f"Sheet: {sheet_name}\n"
        f"Question: {question}\n"
        f"Structured data: {pairs_to_compact_json(pairs)}"
    )
    completion = client.chat.completions.create(
        model=model,
        messages=[
            {"role": "system", "content": "You produce concise and faithful clinical English."},
            {"role": "user", "content": prompt},
        ],
    )
    content = completion.choices[0].message.content or ""
    return normalize_text(content)


def safe_generate_answer(
    client: OpenAI | None,
    model: str,
    question: str,
    sheet_name: str,
    pairs: List[Tuple[str, str]],
    fallback: str,
) -> str:
    if not pairs:
        return ""
    if client is None:
        return fallback
    ans = call_llm_answer(
        client=client,
        model=model,
        question=question,
        sheet_name=sheet_name,
        pairs=pairs,
    )
    return ans if ans else fallback


def process_one_patient(
    item: Dict[str, Any],
    uid: str,
    output_path: Path,
    overwrite: bool,
    client: OpenAI | None,
    model: str,
) -> Dict[str, Any]:
    if output_path.exists() and not overwrite:
        return {"uid": uid, "skipped": True, "status": "exists"}

    record = json_record_template(
        uid=uid,
        group=item["group"],
        patient_name=item["patient_name"],
        patient_dir=item["patient_dir"],
    )
    record["meta"]["model"] = model if client else "fallback_rule_based"

    textrecord_path: Optional[Path] = item["textrecord"]
    if textrecord_path is None or not textrecord_path.exists():
        output_path.write_text(json.dumps(record, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"uid": uid, "skipped": False, "status": "missing_textrecord"}

    wb = load_workbook(filename=str(textrecord_path), data_only=True)

    basic_sheet = get_sheet(wb, "basicinfo")
    if basic_sheet is not None:
        basic_pairs = sheet_to_pairs(basic_sheet)
        basic_answers = build_basic_info_answers(basic_sheet)
        if basic_answers.get("basic_info"):
            basic_subset: List[Tuple[str, str]] = []
            age = extract_field(basic_pairs, DOB_KEYWORDS)
            gender = extract_field(basic_pairs, ["gender", "sex"])
            height = extract_field(basic_pairs, ["height"])
            weight = extract_field(basic_pairs, ["weight"])
            if age:
                basic_subset.append(("DOB", age))
            if gender:
                basic_subset.append(("gender", gender))
            if height:
                basic_subset.append(("height", height))
            if weight:
                basic_subset.append(("weight", weight))
            answer = safe_generate_answer(
                client=client,
                model=model,
                question=QUESTION_BASIC_INFO,
                sheet_name="basicinfo",
                pairs=basic_subset,
                fallback=basic_answers["basic_info"],
            )
            add_task_result(record, "basic_info", QUESTION_BASIC_INFO, textrecord_path, answer)
        if basic_answers.get("primary_concern"):
            cc = extract_field(basic_pairs, ["cc", "chief complaint", "chiefconcern", "primary concern"])
            answer = safe_generate_answer(
                client=client,
                model=model,
                question=QUESTION_PRIMARY_CONCERN,
                sheet_name="basicinfo",
                pairs=[("cc", cc)] if cc else [],
                fallback=basic_answers["primary_concern"],
            )
            add_task_result(
                record,
                "primary_concern",
                QUESTION_PRIMARY_CONCERN,
                textrecord_path,
                answer,
            )
        if basic_answers.get("past_history"):
            rmh = extract_field(
                basic_pairs,
                ["rmh", "medical history", "surgical history", "past medical", "past surgical"],
            )
            answer = safe_generate_answer(
                client=client,
                model=model,
                question=QUESTION_PAST_HISTORY,
                sheet_name="basicinfo",
                pairs=[("RMH", rmh)] if rmh else [],
                fallback=basic_answers["past_history"],
            )
            add_task_result(
                record,
                "past_medical_history",
                QUESTION_PAST_HISTORY,
                textrecord_path,
                answer,
            )

    facial_sheet = get_sheet(wb, "facial photo")
    if facial_sheet is not None:
        facial_pairs = sheet_to_pairs(facial_sheet)
        ans = safe_generate_answer(
            client=client,
            model=model,
            question=QUESTION_FACIAL,
            sheet_name="facial photo",
            pairs=facial_pairs,
            fallback=build_facial_answer(facial_sheet),
        )
        if ans:
            add_task_result(record, "facial_image", QUESTION_FACIAL, textrecord_path, ans)

    dental_sheet = get_sheet(wb, "dental photo")
    if dental_sheet is not None:
        dental_pairs = sheet_to_pairs(dental_sheet)
        ans = safe_generate_answer(
            client=client,
            model=model,
            question=QUESTION_DENTAL,
            sheet_name="dental photo",
            pairs=dental_pairs,
            fallback=build_dental_answer(dental_sheet),
        )
        if ans:
            add_task_result(record, "intraoral_image", QUESTION_DENTAL, textrecord_path, ans)

    tmj_sheet = get_sheet(wb, "TMJ")
    if tmj_sheet is not None:
        tmj_pairs = tmj_sheet_to_pairs_without_mri(tmj_sheet)
        ans = safe_generate_answer(
            client=client,
            model=model,
            question=QUESTION_TMJ,
            sheet_name="TMJ",
            pairs=tmj_pairs,
            fallback=build_tmj_answer(tmj_sheet),
        )
        if ans:
            add_task_result(record, "tmj_clinical_examination", QUESTION_TMJ, textrecord_path, ans)

        ect_ans = build_ect_answer(tmj_sheet)
        if ect_ans:
            add_task_result(record, "ect_examination", QUESTION_ECT, textrecord_path, ect_ans)

    output_path.write_text(json.dumps(record, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"uid": uid, "skipped": False, "status": "written"}


def run_pipeline(args: argparse.Namespace) -> None:
    dataset_root = Path(args.dataset_root).resolve()
    output_dir = Path(args.output_dir).resolve()
    ensure_dir(output_dir)

    patients = discover_patients(dataset_root)
    if args.max_patients is not None:
        patients = patients[: args.max_patients]

    used_uid: Dict[str, int] = {}
    total = len(patients)
    workers = max(1, args.workers)
    api_key = os.getenv(args.api_key_env, "")
    if (not args.dry_run) and (not api_key):
        raise RuntimeError(f"Missing API key. Set environment variable: {args.api_key_env}")
    client = None if args.dry_run else OpenAI(api_key=api_key, base_url=args.api_base)

    print(
        f"[INFO] Found {total} patients under {dataset_root}; "
        f"workers={workers}, dry_run={args.dry_run}"
    )

    jobs: List[Tuple[Dict[str, Any], str, Path]] = []
    for item in patients:
        uid = build_unique_patient_uid(item["group"], item["patient_name"], item["patient_dir"], used_uid)
        output_path = output_dir / f"{uid}.json"
        jobs.append((item, uid, output_path))

    try:
        from tqdm import tqdm  # type: ignore

        progress_iter: Any = tqdm(total=total, desc="patients", unit="patient")
        use_tqdm = True
    except Exception:
        progress_iter = None
        use_tqdm = False

    completed = 0
    with ThreadPoolExecutor(max_workers=workers) as executor:
        future_map = {
            executor.submit(
                process_one_patient,
                item,
                uid,
                output_path,
                args.overwrite,
                client,
                args.model,
            ): (uid, output_path)
            for item, uid, output_path in jobs
        }
        for future in as_completed(future_map):
            uid, output_path = future_map[future]
            completed += 1
            try:
                res = future.result()
                msg = f"[{completed}/{total}] {uid}: {res['status']} -> {output_path}"
            except Exception as e:
                msg = f"[{completed}/{total}] {uid}: ERROR - {e}"

            if use_tqdm:
                progress_iter.update(1)
                progress_iter.write(msg)
            else:
                print(msg)

    if use_tqdm:
        progress_iter.close()

    print("[INFO] Pipeline completed.")


if __name__ == "__main__":
    run_pipeline(parse_args())
